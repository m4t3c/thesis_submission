"""Lettura dell'elenco laureandi da file xlsx.

I file arrivano da estrazioni diverse dei gestionali di ateneo e non hanno
intestazioni identiche. Le due varianti incontrate finora:

    matricola | CORSO       | cognome | ... | email | ate | cellulare
    p06_cds_des | p04_mat_matricola | p01_anaper_cognome | ... |
        p01_anaper_email | p01_anaper_email_ate | p01_anaper_cellulare

Invece di elencare i nomi esatti (che cambierebbero al prossimo formato), le
colonne si riconoscono da un frammento del nome. Serve che il frammento sia
abbastanza specifico da non colpire altre colonne: "ate" compare solo nelle
due colonne dell'email di ateneo, mai in "matricola", "email" o "cellulare".
"""
import re

# Frammenti cercati nel nome della colonna (confronto in minuscolo).
# L'ordine conta: si usa la prima colonna che corrisponde.
FRAMMENTI_CORSO = ("corso", "cds_des")
FRAMMENTI_EMAIL = ("ate",)
FRAMMENTI_COGNOME = ("cognome",)
FRAMMENTI_NOME = ("nome",)

# Un elenco laureandi e' un file piccolo: oltre questa soglia non e' quello
# che ci si aspetta, e conviene fermarsi prima di leggerlo tutto in memoria.
MAX_BYTE_XLSX = 5 * 1024 * 1024


class ErroreXlsx(Exception):
    """Il file non e' leggibile o non ha la struttura attesa."""


def _indice_colonna(intestazioni, frammenti, escludi=()):
    """Posizione della prima colonna il cui nome contiene uno dei frammenti."""
    for frammento in frammenti:
        for i, nome in enumerate(intestazioni):
            if i in escludi:
                continue
            if frammento in (nome or "").strip().lower():
                return i
    return None


def leggi_elenco(file_caricato):
    """Estrae corso di laurea ed email degli studenti da un xlsx.

    Restituisce (corso, [email, ...]). Solleva ErroreXlsx se il file non e'
    leggibile o se mancano le colonne necessarie.
    """
    # Import locale: openpyxl serve solo qui, e tenerlo fuori dall'avvio
    # dell'applicazione evita di pagarne il caricamento a ogni richiesta.
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dipendenza mancante
        raise ErroreXlsx("Libreria di lettura xlsx non disponibile.") from exc

    try:
        # read_only: non carica l'intero foglio in memoria.
        # data_only: prende i valori calcolati, non le formule.
        libro = load_workbook(file_caricato, read_only=True, data_only=True)
    except Exception as exc:
        raise ErroreXlsx(
            "Impossibile leggere il file: assicurati che sia un .xlsx valido."
        ) from exc

    foglio = libro.worksheets[0]
    righe = foglio.iter_rows(values_only=True)

    try:
        intestazioni = [str(c) if c is not None else "" for c in next(righe)]
    except StopIteration:
        raise ErroreXlsx("Il file è vuoto.")

    i_corso = _indice_colonna(intestazioni, FRAMMENTI_CORSO)
    i_email = _indice_colonna(intestazioni, FRAMMENTI_EMAIL)

    if i_email is None:
        raise ErroreXlsx(
            "Nel file non c'è una colonna con l'email di ateneo degli studenti "
            "(attesa una colonna il cui nome contenga «ate»)."
        )
    if i_corso is None:
        raise ErroreXlsx(
            "Nel file non c'è una colonna con il corso di laurea "
            "(attesa una colonna «corso» o «..._cds_des»)."
        )

    corso = ""
    email = []
    for riga in righe:
        if riga is None:
            continue
        valore_email = _testo(riga, i_email)
        if valore_email and "@" in valore_email:
            email.append(valore_email.lower())
        if not corso:
            corso = _testo(riga, i_corso)

    if not email:
        raise ErroreXlsx("Nel file non è stato trovato nessuno studente.")

    libro.close()
    return corso, _senza_duplicati(email)


def _testo(riga, indice):
    if indice is None or indice >= len(riga):
        return ""
    valore = riga[indice]
    return "" if valore is None else str(valore).strip()


def _senza_duplicati(valori):
    """Elimina i doppioni conservando l'ordine di comparsa nel file."""
    visti = set()
    risultato = []
    for v in valori:
        if v not in visti:
            visti.add(v)
            risultato.append(v)
    return risultato
